import pandas as pd
import openpyxl
from logic import load_source_data, process_data, write_to_template
import os

def verify_yamato():
    source_file = "受注集計表（上書き用）.xlsx"
    template_file = "集計表251117.xlsm"
    output_file = "verify_yamato_output.xlsm"
    
    print("--- 1. Loading Source ---")
    filtered_df, col_mapping = load_source_data(source_file)
    
    print("--- 2. Processing Data ---")
    df_koda, df_yamato, koda_headers, yamato_headers, yamato_delivery_types = process_data(filtered_df, col_mapping)
    
    # Check Yamato Sorting
    print("Checking Yamato headers sorting...")
    # Priority: "ﾗﾐ", "ﾗﾐ(ｻﾈｯﾄ行)", "桜花便", "社内便1", "社内便2", "社内便3"
    priority_list = ["ﾗﾐ", "ﾗﾐ(ｻﾈｯﾄ行)", "桜花便", "社内便1", "社内便2", "社内便3"]
    
    # We need to check if the headers in yamato_headers follow this priority
    # We need to map back headers to delivery types to check
    # Let's create a map from name to type from col_mapping
    name_to_type = {m['customer_name']: m['delivery_type'] for m in col_mapping}
    
    current_priority_idx = -1
    sort_fail = False
    
    for name in yamato_headers:
        dtype = name_to_type.get(name)
        if dtype in priority_list:
            idx = priority_list.index(dtype)
        else:
            idx = 999
            
        if idx < current_priority_idx:
            print(f"FAIL: Sort order violation. {name} ({dtype}) came after previous higher priority.")
            sort_fail = True
        current_priority_idx = idx
        
    if not sort_fail:
        print("PASS: Yamato headers sorted by priority.")
    else:
        print("FAIL: Yamato headers sorting failed.")
        print("Headers:", yamato_headers)
        
    print("--- 3. Writing to Template ---")
    output = write_to_template(template_file, df_koda, df_yamato, koda_headers, yamato_headers, yamato_delivery_types, output_file)
    with open(output_file, "wb") as f:
        f.write(output.read())
        
    print("--- 4. Verifying Output File ---")
    wb = openpyxl.load_workbook(output_file)
    if "ラミヤマトその他" in wb.sheetnames:
        ws = wb["ラミヤマトその他"]
        
        # Check Row 2 (Delivery Types)
        print("Checking Row 2 Delivery Types...")
        row2_vals = []
        for c in range(4, 4 + len(yamato_headers)):
            row2_vals.append(ws.cell(row=2, column=c).value)
        
        if row2_vals == yamato_delivery_types:
            print("PASS: Row 2 delivery types match.")
        else:
            print("FAIL: Row 2 delivery types mismatch.")
            print("Expected:", yamato_delivery_types[:5])
            print("Actual:  ", row2_vals[:5])
            
        # Check Row 3 (Customer Names)
        print("Checking Row 3 Customer Names...")
        row3_vals = []
        for c in range(4, 4 + len(yamato_headers)):
            row3_vals.append(ws.cell(row=3, column=c).value)
        
        if row3_vals == yamato_headers:
            print("PASS: Row 3 customer names match.")
        else:
            print("FAIL: Row 3 customer names mismatch.")
            print("Expected:", yamato_headers[:5])
            print("Actual:  ", row3_vals[:5])
            
        # Check Footer
        # Scan for footer
        footer_row = None
        for r in range(1, 100):
            val = ws.cell(row=r, column=1).value
            if val and "商品コード" in str(val):
                if r > 2: # Assume header is 2
                    footer_row = r
                    break
        
        if footer_row:
            print(f"Detected Footer Row: {footer_row}")
            footer_vals = []
            for c in range(4, 4 + len(yamato_headers)):
                footer_vals.append(ws.cell(row=footer_row, column=c).value)
            if footer_vals == yamato_headers:
                print("PASS: Footer headers match.")
            else:
                print("FAIL: Footer headers mismatch.")
        else:
            print("WARN: Footer row not found.")
    else:
        print("FAIL: Sheet 'ラミヤマトその他' not found.")

if __name__ == "__main__":
    verify_yamato()
