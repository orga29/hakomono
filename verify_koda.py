import pandas as pd
import openpyxl
from logic import load_source_data, process_data, write_to_template
import os

def verify_koda():
    source_file = "受注集計表（上書き用）.xlsx"
    template_file = "集計表251117.xlsm"
    output_file = "verify_koda_output.xlsm"
    
    print("--- 1. Loading Source ---")
    filtered_df, col_mapping = load_source_data(source_file)
    print(f"Filtered rows: {len(filtered_df)}")
    
    # Check exclusion
    print("Checking for excluded codes (>= 900)...")
    excluded_found = False
    for m in col_mapping:
        if m['customer_code'].startswith('9'): # Simple check
            # Double check int
            try:
                if int(m['customer_code'].split('-')[0]) >= 900:
                    print(f"FAIL: Found excluded code {m['customer_code']}")
                    excluded_found = True
            except:
                pass
    if not excluded_found:
        print("PASS: No 900+ codes found.")
        
    # Check sorting (Product Code)
    print("Checking product code sorting...")
    codes = filtered_df['product_code'].tolist()
    if codes == sorted(codes, key=str):
        print("PASS: Product codes sorted.")
    else:
        print("FAIL: Product codes NOT sorted.")
        print(codes[:5])
        
    print("--- 2. Processing Data ---")
    df_koda, df_yamato, koda_headers, yamato_headers, yamato_delivery_types = process_data(filtered_df, col_mapping)
    
    # Check column sorting (Customer Code)
    print("Checking customer column sorting...")
    # We need to know the codes for the headers to verify sorting
    # But logic.py sorts col_mapping before extracting.
    # Let's check if the headers match the sorted mapping
    sorted_mapping = sorted(col_mapping, key=lambda x: x['sort_key'])
    expected_koda_names = [m['customer_name'] for m in sorted_mapping if m['delivery_type'] == 'ﾗﾐ']
    
    if koda_headers == expected_koda_names:
        print("PASS: Koda headers sorted by customer code.")
    else:
        print("FAIL: Koda headers NOT sorted correctly.")
        print("Actual:", koda_headers[:3])
        print("Expected:", expected_koda_names[:3])
        
    print("--- 3. Writing to Template ---")
    output = write_to_template(template_file, df_koda, df_yamato, koda_headers, yamato_headers, yamato_delivery_types, output_file)
    with open(output_file, "wb") as f:
        f.write(output.read())
        
    print("--- 4. Verifying Output File ---")
    wb = openpyxl.load_workbook(output_file)
    ws = wb["ラミ（こだ）"]
    
    # Check Header (Row 2)
    print("Checking Row 2 Headers...")
    row2_vals = []
    for c in range(4, 4 + len(koda_headers)):
        row2_vals.append(ws.cell(row=2, column=c).value)
    
    if row2_vals == koda_headers:
        print("PASS: Row 2 headers match.")
    else:
        print("FAIL: Row 2 headers mismatch.")
        print(row2_vals[:3])
        
    # Check Footer Header (Row 15 assumed, or dynamic)
    # We need to find where it wrote.
    # Logic scans for "商品コード" in A.
    # Let's scan A again to find footer.
    footer_row = None
    header_row = None
    for r in range(1, 100):
        val = ws.cell(row=r, column=1).value
        if val and "商品コード" in str(val):
            if header_row is None:
                header_row = r
            else:
                footer_row = r
                break
                
    print(f"Detected Header Row: {header_row}, Footer Row: {footer_row}")
    
    if footer_row:
        print("Checking Footer Headers...")
        footer_vals = []
        for c in range(4, 4 + len(koda_headers)):
            footer_vals.append(ws.cell(row=footer_row, column=c).value)
        if footer_vals == koda_headers:
            print("PASS: Footer headers match.")
        else:
            print("FAIL: Footer headers mismatch.")
    else:
        print("WARN: Footer row not found (might be overwritten or logic error).")

if __name__ == "__main__":
    verify_koda()
