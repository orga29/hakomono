from copy import copy
from datetime import datetime, timedelta
import io
from pathlib import Path
import re
from urllib.parse import parse_qs, urlparse
import zipfile
from zoneinfo import ZoneInfo

import openpyxl
import pandas as pd
from openpyxl.workbook.properties import CalcProperties
from openpyxl.formula.translate import Translator
from openpyxl.formula.translate import TranslatorError
from openpyxl.styles import Border, Side
from openpyxl.utils import get_column_letter


DEFAULT_SOURCE_SHEET_URL = (
    "https://docs.google.com/spreadsheets/d/"
    "19m_l9Wi6P0iDRLiTLJd5wSAvbp4B9Ls3bapFQDOXQrw/edit?gid=558391707#gid=558391707"
)
DEFAULT_TEMPLATE_PATH = Path(__file__).with_name("hakomono_template.xlsx")

BOX_LABEL = "箱"
PRODUCT_CODE_LABEL = "商品コード"
PRODUCT_NAME_LABEL = "商品名"
BOX_TYPE_LABEL = "箱/こ/不"
KODA_SHEET_NAME = "ラミ（こだ）"
YAMATO_SHEET_NAME = "ラミヤマトその他"
KODA_KEYWORD = "こだ"
YAMATO_PRIORITY = [
    "ﾗﾐ",
    "ﾗﾐ(ｻﾈｯﾄ行)",
    "桜花便",
    "社内便1",
    "社内便2",
    "社内便3",
]

KODA_HEADER_ROW = 2
KODA_DATA_START_ROW = 3
KODA_SAMPLE_DATA_END_ROW = 14
KODA_FOOTER_HEADER_ROW = 15
KODA_FOOTER_TOTAL_ROW = 16
KODA_FOOTER_KOMONO_ROW = 17
KODA_FOOTER_GRAND_TOTAL_ROW = 18
KODA_SAMPLE_FOOTER_END_ROW = 18
KODA_CUSTOMER_START_COL = 4
KODA_SAMPLE_CUSTOMER_END_COL = 48  # AV
KODA_SAMPLE_TOTAL_COL = 49  # AW

YAMATO_DELIVERY_ROW = 2
YAMATO_HEADER_ROW = 3
YAMATO_DATA_START_ROW = 4
YAMATO_SAMPLE_DATA_END_ROW = 37
YAMATO_FOOTER_HEADER_ROW = 38
YAMATO_FOOTER_TOTAL_ROW = 39
YAMATO_FOOTER_KOMONO_ROW = 40
YAMATO_FOOTER_GRAND_TOTAL_ROW = 41
YAMATO_SAMPLE_FOOTER_END_ROW = 41
YAMATO_CUSTOMER_START_COL = 4
YAMATO_SAMPLE_CUSTOMER_END_COL = 31  # AE
YAMATO_SAMPLE_TOTAL_COL = 32  # AF
YAMATO_UNIFORM_WIDTH = 5.57
YAMATO_HEADER_HEIGHT = 58.5
YAMATO_SUMMARY_HEIGHT = 24.75


def suggested_output_filename(now=None):
    now = now or datetime.now(ZoneInfo("Asia/Tokyo"))
    return f"hakomono-{now.strftime('%m%d')}.xlsx"


def load_source_data(source_url=DEFAULT_SOURCE_SHEET_URL):
    df_raw = _read_source_dataframe(source_url)
    code_row_idx, delivery_row_idx, customer_row_idx, data_start_idx = _detect_source_layout(df_raw)

    customer_codes = df_raw.iloc[code_row_idx, 3:]
    delivery_types = df_raw.iloc[delivery_row_idx, 3:]
    customer_names = df_raw.iloc[customer_row_idx, 3:]

    col_mapping = []
    for idx, (code, delivery_type, customer_name) in enumerate(
        zip(customer_codes, delivery_types, customer_names)
    ):
        code_str = _normalize_text(code)
        delivery_type_str = _normalize_text(delivery_type)
        customer_name_str = _normalize_text(customer_name)

        if not code_str or not customer_name_str or _is_excluded_customer_code(code_str):
            continue

        col_mapping.append(
            {
                "col_idx": idx + 3,
                "customer_code": code_str,
                "delivery_type": delivery_type_str,
                "customer_name": customer_name_str,
                "sort_key": _get_sort_key(code_str),
            }
        )

    data_df = df_raw.iloc[data_start_idx:].copy()
    data_df = data_df.rename(columns={0: "product_code", 1: "product_name", 2: "box_type"})
    data_df["box_type"] = data_df["box_type"].map(_normalize_text)
    filtered_df = data_df[data_df["box_type"] == BOX_LABEL].copy()

    filtered_df["product_code"] = filtered_df["product_code"].map(_normalize_text)
    filtered_df["product_name"] = filtered_df["product_name"].map(_normalize_text)
    filtered_df["_product_sort_key"] = filtered_df["product_code"].map(_get_sort_key)
    filtered_df = filtered_df.sort_values(
        by=["_product_sort_key", "product_code"], kind="stable"
    ).drop(columns=["_product_sort_key"])

    return filtered_df, col_mapping


def process_data(filtered_df, col_mapping):
    base_cols = filtered_df.loc[:, ["product_code", "product_name"]].copy()
    base_cols.columns = [PRODUCT_CODE_LABEL, PRODUCT_NAME_LABEL]
    base_cols[BOX_TYPE_LABEL] = BOX_LABEL

    koda_mapping = []
    yamato_mapping = []
    for item in col_mapping:
        if _is_koda_delivery_type(item["delivery_type"]):
            koda_mapping.append(item)
        else:
            yamato_mapping.append(item)

    koda_mapping.sort(key=lambda item: item["sort_key"])
    yamato_mapping.sort(key=_yamato_sort_key)

    def build_df_and_headers(mapping, include_delivery_types=False):
        quantity_cols = []
        headers = []
        delivery_types = []

        for item in mapping:
            col_data = filtered_df.iloc[:, item["col_idx"]].copy()
            col_data.name = item["customer_name"]
            quantity_cols.append(col_data)
            headers.append(item["customer_name"])
            if include_delivery_types:
                delivery_types.append(item["delivery_type"])

        df = pd.concat([base_cols] + quantity_cols, axis=1) if quantity_cols else base_cols.copy()
        if include_delivery_types:
            return df, headers, delivery_types
        return df, headers

    df_koda, koda_headers = build_df_and_headers(koda_mapping)
    df_yamato, yamato_headers, yamato_delivery_types = build_df_and_headers(
        yamato_mapping, include_delivery_types=True
    )

    return df_koda, df_yamato, koda_headers, yamato_headers, yamato_delivery_types


def write_to_template(
    template_path,
    df_koda,
    df_yamato,
    koda_headers,
    yamato_headers,
    yamato_delivery_types,
):
    template_path = Path(template_path)
    if not template_path.exists():
        raise FileNotFoundError(f"テンプレートが見つかりません: {template_path}")

    workbook = openpyxl.load_workbook(template_path)

    if KODA_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"テンプレートに '{KODA_SHEET_NAME}' シートがありません。")
    if YAMATO_SHEET_NAME not in workbook.sheetnames:
        raise ValueError(f"テンプレートに '{YAMATO_SHEET_NAME}' シートがありません。")

    _write_koda_sheet(workbook[KODA_SHEET_NAME], df_koda, koda_headers)
    _write_yamato_sheet(
        workbook[YAMATO_SHEET_NAME],
        df_yamato,
        yamato_headers,
        yamato_delivery_types,
    )
    workbook.calculation = CalcProperties(
        calcMode="auto",
        fullCalcOnLoad=True,
        forceFullCalc=True,
    )

    output = io.BytesIO()
    workbook.save(output)
    output.seek(0)
    return _restore_template_drawings(template_path, output)


def _read_source_dataframe(source_url):
    local_path = Path(source_url)
    if local_path.exists():
        if local_path.suffix.lower() in {".xlsx", ".xlsm", ".xls"}:
            df = pd.read_excel(local_path, sheet_name=0, header=None)
        else:
            df = pd.read_csv(local_path, header=None)
        return df.dropna(axis=0, how="all").dropna(axis=1, how="all")

    export_url = _build_google_sheet_csv_export_url(source_url)
    df = pd.read_csv(export_url, header=None)
    return df.dropna(axis=0, how="all").dropna(axis=1, how="all")


def _detect_source_layout(df_raw):
    max_scan_rows = min(30, len(df_raw))
    for row_idx in range(max_scan_rows):
        col_a = _normalize_text(df_raw.iloc[row_idx, 0] if df_raw.shape[1] > 0 else "")
        col_b = _normalize_text(df_raw.iloc[row_idx, 1] if df_raw.shape[1] > 1 else "")
        col_c = _normalize_text(df_raw.iloc[row_idx, 2] if df_raw.shape[1] > 2 else "")

        if col_a == PRODUCT_CODE_LABEL and col_b == PRODUCT_NAME_LABEL and BOX_LABEL in col_c:
            if row_idx < 2:
                raise ValueError("元データのヘッダー位置が想定より上にあり、配送先情報を特定できません。")
            return row_idx - 2, row_idx - 1, row_idx, row_idx + 1

    return 5, 6, 7, 8


def _extract_sheet_id(sheet_url):
    match = re.search(r"/spreadsheets/d/([a-zA-Z0-9-_]+)", sheet_url)
    if not match:
        raise ValueError("Google スプレッドシート URL からシート ID を取得できませんでした。")
    return match.group(1)


def _extract_gid(sheet_url):
    parsed = urlparse(sheet_url)
    gid = parse_qs(parsed.query).get("gid", [None])[0]
    if gid is None and parsed.fragment.startswith("gid="):
        gid = parsed.fragment.split("=", 1)[1]
    return gid or "0"


def _build_google_sheet_csv_export_url(sheet_url):
    sheet_id = _extract_sheet_id(sheet_url)
    gid = _extract_gid(sheet_url)
    return f"https://docs.google.com/spreadsheets/d/{sheet_id}/export?format=csv&gid={gid}"


def _normalize_text(value):
    if pd.isna(value):
        return ""
    return str(value).strip()


def _is_excluded_customer_code(code_str):
    match = re.match(r"^(\d+)", code_str)
    return bool(match and int(match.group(1)) >= 900)


def _get_sort_key(text):
    parts = re.findall(r"\d+", _normalize_text(text))
    if not parts:
        return (float("inf"), _normalize_text(text))
    return tuple(int(part) for part in parts)


def _is_koda_delivery_type(delivery_type):
    return KODA_KEYWORD in _normalize_text(delivery_type)


def _yamato_sort_key(item):
    delivery_type = _normalize_text(item["delivery_type"])
    priority_index = len(YAMATO_PRIORITY)
    for idx, keyword in enumerate(YAMATO_PRIORITY):
        if delivery_type == keyword:
            priority_index = idx
            break
    if priority_index == len(YAMATO_PRIORITY):
        for idx, keyword in enumerate(YAMATO_PRIORITY):
            if keyword in delivery_type:
                priority_index = idx
                break
    return (priority_index, item["sort_key"], item["customer_name"])


def _write_koda_sheet(worksheet, df, customer_names):
    df = _filter_zero_total_rows(df)
    sample_customer_count = KODA_SAMPLE_CUSTOMER_END_COL - KODA_CUSTOMER_START_COL + 1
    desired_customer_count = len(customer_names)

    hidden_start_col = None
    if desired_customer_count > sample_customer_count:
        extra_cols = desired_customer_count - sample_customer_count
        worksheet.insert_cols(KODA_SAMPLE_TOTAL_COL, amount=extra_cols)
        for offset in range(extra_cols):
            _copy_column_style(
                worksheet,
                source_col=KODA_SAMPLE_CUSTOMER_END_COL,
                target_col=KODA_SAMPLE_TOTAL_COL + offset,
                start_row=1,
                end_row=KODA_FOOTER_GRAND_TOTAL_ROW,
            )

    total_col = KODA_CUSTOMER_START_COL + desired_customer_count
    if desired_customer_count <= sample_customer_count:
        if total_col != KODA_SAMPLE_TOTAL_COL:
            _copy_column_style(
                worksheet,
                source_col=KODA_SAMPLE_TOTAL_COL,
                target_col=total_col,
                start_row=1,
                end_row=KODA_FOOTER_GRAND_TOTAL_ROW,
            )
        hidden_start_col = total_col + 1
    _normalize_koda_column_widths(worksheet, total_col)
    _set_hidden_columns(
        worksheet,
        start_col=KODA_CUSTOMER_START_COL,
        end_col=max(total_col, KODA_SAMPLE_TOTAL_COL),
        visible_end_col=total_col,
    )
    data_row_count = len(df)
    sample_data_row_count = KODA_SAMPLE_DATA_END_ROW - KODA_DATA_START_ROW + 1

    odd_template = _capture_row_template(worksheet, KODA_DATA_START_ROW, total_col)
    even_template = _capture_row_template(worksheet, KODA_DATA_START_ROW + 1, total_col)
    last_template = _capture_row_template(worksheet, KODA_SAMPLE_DATA_END_ROW, total_col)
    footer_templates = [
        _capture_row_template(worksheet, KODA_FOOTER_HEADER_ROW + offset, total_col)
        for offset in range(4)
    ]

    if data_row_count > sample_data_row_count:
        worksheet.insert_rows(KODA_FOOTER_HEADER_ROW, amount=data_row_count - sample_data_row_count)

    footer_header_row = KODA_DATA_START_ROW + data_row_count
    footer_total_row = footer_header_row + 1
    footer_komono_row = footer_header_row + 2
    footer_grand_total_row = footer_header_row + 3
    _set_hidden_rows(
        worksheet,
        start_row=KODA_DATA_START_ROW,
        end_row=max(KODA_SAMPLE_FOOTER_END_ROW, footer_grand_total_row),
        visible_end_row=footer_grand_total_row,
    )

    for row_idx in range(KODA_DATA_START_ROW, footer_header_row):
        template = odd_template if (row_idx - KODA_DATA_START_ROW) % 2 == 0 else even_template
        _apply_row_template(worksheet, row_idx, template)
    if data_row_count:
        _apply_bottom_border_from_template(worksheet, last_template, footer_header_row - 1, total_col)

    for offset, template in enumerate(footer_templates):
        _apply_row_template(worksheet, footer_header_row + offset, template)

    _clear_range(worksheet, KODA_HEADER_ROW, KODA_CUSTOMER_START_COL, total_col)
    for row_idx in range(KODA_DATA_START_ROW, footer_header_row):
        _clear_range(worksheet, row_idx, 1, total_col)
    for row_idx in range(footer_header_row, footer_grand_total_row + 1):
        _clear_range(worksheet, row_idx, KODA_CUSTOMER_START_COL, total_col)

    for idx, customer_name in enumerate(customer_names):
        col_idx = KODA_CUSTOMER_START_COL + idx
        display_name = _format_koda_customer_name(customer_name)
        worksheet.cell(KODA_HEADER_ROW, col_idx, display_name)
        worksheet.cell(footer_header_row, col_idx, display_name)

    for row_idx, row in enumerate(df.itertuples(index=False), start=KODA_DATA_START_ROW):
        worksheet.cell(row_idx, 1, row[0])
        worksheet.cell(row_idx, 2, row[1])
        worksheet.cell(row_idx, 3, row[2])

        last_customer_col = total_col - 1
        for value_idx, value in enumerate(row[3:], start=KODA_CUSTOMER_START_COL):
            worksheet.cell(row_idx, value_idx, _clean_excel_value(value))

        if last_customer_col >= KODA_CUSTOMER_START_COL:
            start_ref = f"{get_column_letter(KODA_CUSTOMER_START_COL)}{row_idx}"
            end_ref = f"{get_column_letter(last_customer_col)}{row_idx}"
            worksheet.cell(row_idx, total_col, f"=SUM({start_ref}:{end_ref})")
        else:
            worksheet.cell(row_idx, total_col, 0)

    if data_row_count:
        data_end_row = footer_header_row - 1
        for col_idx in range(KODA_CUSTOMER_START_COL, total_col + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.cell(
                footer_total_row,
                col_idx,
                f"=SUM({col_letter}{KODA_DATA_START_ROW}:{col_letter}{data_end_row})",
            )
    else:
        for col_idx in range(KODA_CUSTOMER_START_COL, total_col + 1):
            worksheet.cell(footer_total_row, col_idx, 0)

    _clear_range(worksheet, footer_grand_total_row, KODA_CUSTOMER_START_COL, total_col)

    _set_koda_title(worksheet)
    _clear_range(worksheet, KODA_HEADER_ROW, total_col + 1, worksheet.max_column)
    for row_idx in range(KODA_DATA_START_ROW, footer_grand_total_row + 1):
        _clear_range(worksheet, row_idx, total_col + 1, worksheet.max_column)


def _write_yamato_sheet(worksheet, df, customer_names, delivery_types):
    df = _filter_zero_total_rows(df)
    sample_customer_count = YAMATO_SAMPLE_CUSTOMER_END_COL - YAMATO_CUSTOMER_START_COL + 1
    desired_customer_count = len(customer_names)

    if desired_customer_count > sample_customer_count:
        extra_cols = desired_customer_count - sample_customer_count
        worksheet.insert_cols(YAMATO_SAMPLE_TOTAL_COL, amount=extra_cols)
        for offset in range(extra_cols):
            _copy_column_style(
                worksheet,
                source_col=YAMATO_SAMPLE_CUSTOMER_END_COL,
                target_col=YAMATO_SAMPLE_TOTAL_COL + offset,
                start_row=1,
                end_row=YAMATO_FOOTER_GRAND_TOTAL_ROW,
            )

    total_col = YAMATO_CUSTOMER_START_COL + desired_customer_count
    if desired_customer_count <= sample_customer_count and total_col != YAMATO_SAMPLE_TOTAL_COL:
        _copy_column_style(
            worksheet,
            source_col=YAMATO_SAMPLE_TOTAL_COL,
            target_col=total_col,
            start_row=1,
            end_row=YAMATO_FOOTER_GRAND_TOTAL_ROW,
        )

    _normalize_yamato_column_widths(worksheet, total_col)
    _set_hidden_columns(
        worksheet,
        start_col=YAMATO_CUSTOMER_START_COL,
        end_col=max(total_col, YAMATO_SAMPLE_TOTAL_COL),
        visible_end_col=total_col,
    )

    data_row_count = len(df)
    sample_data_row_count = YAMATO_SAMPLE_DATA_END_ROW - YAMATO_DATA_START_ROW + 1

    odd_template = _capture_row_template(worksheet, YAMATO_DATA_START_ROW, total_col)
    even_template = _capture_row_template(worksheet, YAMATO_DATA_START_ROW + 1, total_col)
    last_template = _capture_row_template(worksheet, YAMATO_SAMPLE_DATA_END_ROW, total_col)
    footer_templates = [
        _capture_row_template(worksheet, YAMATO_FOOTER_HEADER_ROW + offset, total_col)
        for offset in range(4)
    ]

    if data_row_count > sample_data_row_count:
        worksheet.insert_rows(YAMATO_FOOTER_HEADER_ROW, amount=data_row_count - sample_data_row_count)

    footer_header_row = YAMATO_DATA_START_ROW + data_row_count
    footer_total_row = footer_header_row + 1
    footer_komono_row = footer_header_row + 2
    footer_grand_total_row = footer_header_row + 3

    _set_hidden_rows(
        worksheet,
        start_row=YAMATO_DATA_START_ROW,
        end_row=max(YAMATO_SAMPLE_FOOTER_END_ROW, footer_grand_total_row),
        visible_end_row=footer_grand_total_row,
    )

    for row_idx in range(YAMATO_DATA_START_ROW, footer_header_row):
        template = odd_template if (row_idx - YAMATO_DATA_START_ROW) % 2 == 0 else even_template
        _apply_row_template(worksheet, row_idx, template)
    if data_row_count:
        _apply_bottom_border_from_template(worksheet, last_template, footer_header_row - 1, total_col)

    for offset, template in enumerate(footer_templates):
        _apply_row_template(worksheet, footer_header_row + offset, template)

    worksheet.row_dimensions[YAMATO_HEADER_ROW].height = YAMATO_HEADER_HEIGHT
    worksheet.row_dimensions[footer_header_row].height = YAMATO_HEADER_HEIGHT
    worksheet.row_dimensions[footer_total_row].height = YAMATO_SUMMARY_HEIGHT
    worksheet.row_dimensions[footer_komono_row].height = YAMATO_SUMMARY_HEIGHT
    worksheet.row_dimensions[footer_grand_total_row].height = YAMATO_SUMMARY_HEIGHT

    _clear_range(worksheet, YAMATO_DELIVERY_ROW, YAMATO_CUSTOMER_START_COL, total_col)
    _clear_range(worksheet, YAMATO_HEADER_ROW, YAMATO_CUSTOMER_START_COL, total_col)
    _clear_range(worksheet, footer_header_row, YAMATO_CUSTOMER_START_COL, total_col)
    for row_idx in range(YAMATO_DATA_START_ROW, footer_header_row):
        _clear_range(worksheet, row_idx, 1, total_col)
    for row_idx in range(footer_total_row, footer_grand_total_row + 1):
        _clear_range(worksheet, row_idx, YAMATO_CUSTOMER_START_COL, total_col)

    for idx, (delivery_type, customer_name) in enumerate(zip(delivery_types, customer_names)):
        col_idx = YAMATO_CUSTOMER_START_COL + idx
        worksheet.cell(YAMATO_DELIVERY_ROW, col_idx, delivery_type)
        worksheet.cell(YAMATO_HEADER_ROW, col_idx, customer_name)
        worksheet.cell(footer_header_row, col_idx, customer_name)

    for row_idx, row in enumerate(df.itertuples(index=False), start=YAMATO_DATA_START_ROW):
        worksheet.cell(row_idx, 1, row[0])
        worksheet.cell(row_idx, 2, row[1])
        worksheet.cell(row_idx, 3, row[2])
        last_customer_col = total_col - 1
        for value_idx, value in enumerate(row[3:], start=YAMATO_CUSTOMER_START_COL):
            worksheet.cell(row_idx, value_idx, _clean_excel_value(value))
        if last_customer_col >= YAMATO_CUSTOMER_START_COL:
            start_ref = f"{get_column_letter(YAMATO_CUSTOMER_START_COL)}{row_idx}"
            end_ref = f"{get_column_letter(last_customer_col)}{row_idx}"
            worksheet.cell(row_idx, total_col, f"=SUM({start_ref}:{end_ref})")
        else:
            worksheet.cell(row_idx, total_col, 0)

    if data_row_count:
        data_end_row = footer_header_row - 1
        for col_idx in range(YAMATO_CUSTOMER_START_COL, total_col + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.cell(
                footer_total_row,
                col_idx,
                f"=SUM({col_letter}{YAMATO_DATA_START_ROW}:{col_letter}{data_end_row})",
            )
    else:
        for col_idx in range(YAMATO_CUSTOMER_START_COL, total_col + 1):
            worksheet.cell(footer_total_row, col_idx, 0)

    _clear_range(worksheet, footer_grand_total_row, YAMATO_CUSTOMER_START_COL, total_col)
    _set_hidden_columns(
        worksheet,
        start_col=total_col + 1,
        end_col=worksheet.max_column,
        visible_end_col=total_col,
    )
    _set_yamato_group_borders(worksheet, delivery_types, total_col, footer_grand_total_row)
    _set_yamato_title(worksheet)


def _find_header_and_footer_rows(worksheet, default_header, default_footer):
    header_row = None
    footer_row = None

    for row_idx in range(1, min(200, worksheet.max_row) + 1):
        value_a = _normalize_text(worksheet.cell(row=row_idx, column=1).value)
        value_b = _normalize_text(worksheet.cell(row=row_idx, column=2).value)
        if value_a == PRODUCT_CODE_LABEL and value_b == PRODUCT_NAME_LABEL:
            if header_row is None:
                header_row = row_idx
            else:
                footer_row = row_idx
                break

    return header_row or default_header, footer_row or default_footer


def _find_formula_start_column(worksheet, header_row, footer_row):
    scan_row = header_row + 1 if header_row + 1 < footer_row else header_row
    for col_idx in range(4, worksheet.max_column + 1):
        value = worksheet.cell(row=scan_row, column=col_idx).value
        if isinstance(value, str) and value.startswith("="):
            return col_idx
    return worksheet.max_column + 1


def _write_sheet_data(
    worksheet,
    df,
    customer_names,
    header_row,
    footer_row,
    limit_col,
    delivery_type_row=None,
    delivery_types=None,
):
    max_customer_cols = max(0, limit_col - 3)
    if len(customer_names) > max_customer_cols:
        raise ValueError(
            f"{worksheet.title} のテンプレート幅が不足しています。 "
            f"得意先列 {len(customer_names)} 件に対して {max_customer_cols} 件までしか配置できません。"
        )

    data_start_row = header_row + 1
    template_rows = _capture_row_templates(worksheet, data_start_row, footer_row - 1, worksheet.max_column)
    if not template_rows:
        raise ValueError(f"{worksheet.title} のテンプレート行を取得できませんでした。")

    num_data_rows = len(df)
    available_rows = max(0, footer_row - header_row - 1)
    if num_data_rows > available_rows:
        rows_to_insert = num_data_rows - available_rows
        worksheet.insert_rows(footer_row, amount=rows_to_insert)
        footer_row += rows_to_insert
    elif num_data_rows < available_rows:
        worksheet.delete_rows(data_start_row + num_data_rows, available_rows - num_data_rows)
        footer_row -= available_rows - num_data_rows

    if delivery_type_row is not None:
        _clear_range(worksheet, delivery_type_row, 4, limit_col - 1)

    _clear_range(worksheet, header_row, 4, limit_col - 1)
    _clear_range(worksheet, footer_row, 4, limit_col - 1)

    for row_idx in range(data_start_row, footer_row):
        template = template_rows[(row_idx - data_start_row) % len(template_rows)]
        _apply_row_template(worksheet, row_idx, template)
        _clear_range(worksheet, row_idx, 1, limit_col - 1)

    if delivery_type_row is not None and delivery_types is not None:
        for idx, delivery_type in enumerate(delivery_types):
            worksheet.cell(row=delivery_type_row, column=4 + idx, value=delivery_type)

    for idx, customer_name in enumerate(customer_names):
        target_col = 4 + idx
        worksheet.cell(row=header_row, column=target_col, value=customer_name)
        worksheet.cell(row=footer_row, column=target_col, value=customer_name)

    for df_row_idx, row in enumerate(df.itertuples(index=False), start=1):
        current_row = header_row + df_row_idx
        worksheet.cell(row=current_row, column=1, value=row[0])
        worksheet.cell(row=current_row, column=2, value=row[1])
        worksheet.cell(row=current_row, column=3, value=row[2])
        for value_idx, value in enumerate(row[3:], start=4):
            worksheet.cell(row=current_row, column=value_idx, value=_clean_excel_value(value))

    _rebuild_generic_footer_formulas(
        worksheet=worksheet,
        header_row=header_row,
        footer_row=footer_row,
        total_col=limit_col,
    )


def _capture_row_templates(worksheet, start_row, end_row, max_col):
    return [_capture_row_template(worksheet, row_idx, max_col) for row_idx in range(start_row, end_row + 1)]


def _capture_row_template(worksheet, row_idx, max_col):
    template = {
        "source_row": row_idx,
        "height": worksheet.row_dimensions[row_idx].height,
        "hidden": worksheet.row_dimensions[row_idx].hidden,
        "cells": [],
    }
    for col_idx in range(1, max_col + 1):
        cell = worksheet.cell(row=row_idx, column=col_idx)
        template["cells"].append(
            {
                "row": row_idx,
                "col": col_idx,
                "value": cell.value,
                "style": copy(cell._style),
                "number_format": cell.number_format,
                "font": copy(cell.font),
                "fill": copy(cell.fill),
                "border": copy(cell.border),
                "alignment": copy(cell.alignment),
                "protection": copy(cell.protection),
            }
        )
    return template


def _apply_row_template(worksheet, target_row, template):
    row_dimension = worksheet.row_dimensions[target_row]
    row_dimension.height = template["height"]
    row_dimension.hidden = template["hidden"]

    for cell_template in template["cells"]:
        target_cell = worksheet.cell(row=target_row, column=cell_template["col"])
        target_cell._style = copy(cell_template["style"])
        target_cell.number_format = cell_template["number_format"]
        target_cell.font = copy(cell_template["font"])
        target_cell.fill = copy(cell_template["fill"])
        target_cell.border = copy(cell_template["border"])
        target_cell.alignment = copy(cell_template["alignment"])
        target_cell.protection = copy(cell_template["protection"])
        target_cell.value = _translate_formula_if_needed(
            cell_template["value"],
            source_row=cell_template["row"],
            source_col=cell_template["col"],
            target_row=target_row,
            target_col=cell_template["col"],
        )


def _apply_bottom_border_from_template(worksheet, template, target_row, max_col):
    for cell_template in template["cells"]:
        if cell_template["col"] > max_col:
            continue
        target_cell = worksheet.cell(target_row, cell_template["col"])
        target_cell.border = copy(cell_template["border"])


def _copy_column_style(worksheet, source_col, target_col, start_row, end_row):
    source_letter = get_column_letter(source_col)
    target_letter = get_column_letter(target_col)
    worksheet.column_dimensions[target_letter].width = worksheet.column_dimensions[source_letter].width
    worksheet.column_dimensions[target_letter].hidden = worksheet.column_dimensions[source_letter].hidden

    for row_idx in range(start_row, end_row + 1):
        source_cell = worksheet.cell(row_idx, source_col)
        target_cell = worksheet.cell(row_idx, target_col)
        target_cell._style = copy(source_cell._style)
        target_cell.number_format = source_cell.number_format
        target_cell.font = copy(source_cell.font)
        target_cell.fill = copy(source_cell.fill)
        target_cell.border = copy(source_cell.border)
        target_cell.alignment = copy(source_cell.alignment)
        target_cell.protection = copy(source_cell.protection)
        target_cell.value = _translate_formula_if_needed(
            source_cell.value,
            source_row=row_idx,
            source_col=source_col,
            target_row=row_idx,
            target_col=target_col,
        )


def _set_hidden_columns(worksheet, start_col, end_col, visible_end_col):
    for col_idx in range(start_col, end_col + 1):
        letter = get_column_letter(col_idx)
        worksheet.column_dimensions[letter].hidden = col_idx > visible_end_col


def _set_hidden_rows(worksheet, start_row, end_row, visible_end_row):
    for row_idx in range(start_row, end_row + 1):
        worksheet.row_dimensions[row_idx].hidden = row_idx > visible_end_row


def _filter_zero_total_rows(df):
    if df.empty or len(df.columns) <= 3:
        return df.copy()

    quantity_df = df.iloc[:, 3:].apply(pd.to_numeric, errors="coerce").fillna(0)
    mask = quantity_df.sum(axis=1) != 0
    return df.loc[mask].reset_index(drop=True)


def _normalize_koda_column_widths(worksheet, total_col):
    uniform_width = 7
    for col_idx in range(KODA_CUSTOMER_START_COL, total_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = uniform_width


def _format_koda_customer_name(name):
    normalized = _normalize_text(name)
    if normalized.startswith("こだ"):
        return normalized[2:]
    return normalized


def _normalize_yamato_column_widths(worksheet, total_col):
    for col_idx in range(YAMATO_CUSTOMER_START_COL, total_col + 1):
        worksheet.column_dimensions[get_column_letter(col_idx)].width = YAMATO_UNIFORM_WIDTH


def _set_koda_title(worksheet):
    delivery_date = datetime.now(ZoneInfo("Asia/Tokyo")).date() + timedelta(days=1)
    weekdays = "月火水木金土日"
    weekday = weekdays[delivery_date.weekday()]
    worksheet["B1"] = f"{delivery_date.month}/{delivery_date.day}({weekday}) 店着　ラミ便（こだ）"


def _set_yamato_title(worksheet):
    delivery_date = datetime.now(ZoneInfo("Asia/Tokyo")).date() + timedelta(days=1)
    weekdays = "月火水木金土日"
    weekday = weekdays[delivery_date.weekday()]
    worksheet["B1"] = f"{delivery_date.month}/{delivery_date.day}({weekday}) 店着　ラミ便・その他"


def _yamato_group_name(delivery_type):
    normalized = _normalize_text(delivery_type)
    if normalized == "ﾗﾐ":
        return "ﾗﾐ"
    if normalized == "ﾗﾐ(ｻﾈｯﾄ行)":
        return "ﾗﾐ(ｻﾈｯﾄ行)"
    return "その他"


def _set_yamato_group_borders(worksheet, delivery_types, total_col, footer_grand_total_row):
    boundary_cols = []
    previous_group = None
    for idx, delivery_type in enumerate(delivery_types, start=YAMATO_CUSTOMER_START_COL):
        current_group = _yamato_group_name(delivery_type)
        if previous_group is not None and current_group != previous_group:
            boundary_cols.append(idx - 1)
        previous_group = current_group

    medium_side = Side(style="medium")
    for col_idx in boundary_cols:
        for row_idx in range(YAMATO_DELIVERY_ROW, footer_grand_total_row + 1):
            cell = worksheet.cell(row_idx, col_idx)
            current = cell.border
            cell.border = Border(
                left=current.left,
                right=medium_side,
                top=current.top,
                bottom=current.bottom,
                diagonal=current.diagonal,
                diagonal_direction=current.diagonal_direction,
                outline=current.outline,
                vertical=current.vertical,
                horizontal=current.horizontal,
            )


def _translate_formula_if_needed(value, source_row, source_col, target_row, target_col):
    if not (isinstance(value, str) and value.startswith("=")):
        return value
    origin = get_column_letter(source_col) + str(source_row)
    target = get_column_letter(target_col) + str(target_row)
    try:
        return Translator(value, origin=origin).translate_formula(target)
    except TranslatorError:
        return value


def _rebuild_generic_footer_formulas(worksheet, header_row, footer_row, total_col):
    footer_total_row = footer_row + 1
    footer_komono_row = footer_row + 2
    footer_grand_total_row = footer_row + 3
    data_start_row = header_row + 1
    data_end_row = footer_row - 1

    if data_end_row >= data_start_row:
        for col_idx in range(4, total_col + 1):
            col_letter = get_column_letter(col_idx)
            worksheet.cell(
                footer_total_row,
                col_idx,
                f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})",
            )
            worksheet.cell(
                footer_grand_total_row,
                col_idx,
                f"=SUM({col_letter}{footer_total_row}:{col_letter}{footer_komono_row})",
            )
    else:
        for col_idx in range(4, total_col + 1):
            worksheet.cell(footer_total_row, col_idx, 0)
            worksheet.cell(footer_grand_total_row, col_idx, 0)


def _clear_range(worksheet, row_idx, start_col, end_col):
    if end_col < start_col:
        return
    for col_idx in range(start_col, end_col + 1):
        worksheet.cell(row=row_idx, column=col_idx).value = None


def _clean_excel_value(value):
    if pd.isna(value):
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if not stripped:
            return None
        if re.fullmatch(r"-?\d+", stripped):
            return int(stripped)
        if re.fullmatch(r"-?\d+\.\d+", stripped):
            return float(stripped)
    return value


def _restore_template_drawings(template_path, workbook_stream):
    with zipfile.ZipFile(template_path) as template_zip:
        template_entries = {
            name: template_zip.read(name)
            for name in template_zip.namelist()
            if name.startswith("xl/drawings/")
            or name in {
                "xl/worksheets/_rels/sheet1.xml.rels",
                "xl/worksheets/_rels/sheet2.xml.rels",
                "[Content_Types].xml",
            }
        }

    source_bytes = workbook_stream.getvalue()
    output = io.BytesIO()
    with zipfile.ZipFile(io.BytesIO(source_bytes), "r") as src_zip, zipfile.ZipFile(
        output, "w", compression=zipfile.ZIP_DEFLATED
    ) as out_zip:
        existing_names = set(src_zip.namelist())
        template_passthrough_names = {
            name
            for name in template_entries
            if name.startswith("xl/drawings/") or name.startswith("xl/worksheets/_rels/")
        }
        for name in src_zip.namelist():
            if name in template_passthrough_names:
                continue
            data = src_zip.read(name)
            if name in {"xl/worksheets/sheet1.xml", "xl/worksheets/sheet2.xml"}:
                data = _ensure_sheet_drawing_tag(data)
            elif name == "[Content_Types].xml":
                data = _ensure_drawing_content_types(data)
            out_zip.writestr(name, data)

        for name, data in template_entries.items():
            if name in {"[Content_Types].xml"}:
                continue
            if name.startswith("xl/drawings/") or name.startswith("xl/worksheets/_rels/"):
                out_zip.writestr(name, data)

    output.seek(0)
    return output


def _ensure_sheet_drawing_tag(sheet_xml_bytes):
    marker = b'<drawing r:id="rId1"/>'
    updated = sheet_xml_bytes
    if b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"' not in updated:
        updated = updated.replace(
            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main"',
            b'<worksheet xmlns="http://schemas.openxmlformats.org/spreadsheetml/2006/main" '
            b'xmlns:r="http://schemas.openxmlformats.org/officeDocument/2006/relationships"',
            1,
        )
    if marker not in updated:
        updated = updated.replace(b"</worksheet>", marker + b"</worksheet>")
    return updated


def _ensure_drawing_content_types(content_types_bytes):
    required = [
        b'<Override PartName="/xl/drawings/drawing1.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>',
        b'<Override PartName="/xl/drawings/drawing2.xml" ContentType="application/vnd.openxmlformats-officedocument.drawing+xml"/>',
    ]
    updated = content_types_bytes
    for override in required:
        if override not in updated:
            updated = updated.replace(b"</Types>", override + b"</Types>")
    return updated
